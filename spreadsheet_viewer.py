
import csv
import io
import os
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QComboBox,
    QCheckBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QMessageBox,
)

SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".ods", ".csv"}

_BTN_STYLE = """
    QPushButton {
        background: rgba(59,130,246,0.2);
        border: 1px solid rgba(59,130,246,0.4);
        border-radius: 6px;
        padding: 4px 14px;
        color: #60a5fa;
        font-weight: 500;
    }
    QPushButton:hover { background: rgba(59,130,246,0.35); }
"""

_TABLE_STYLE = """
    QTableWidget {
        background: #111827;
        color: #e5e7eb;
        border: 1px solid rgba(148,163,184,0.2);
        border-radius: 8px;
        gridline-color: rgba(148,163,184,0.15);
    }
    QTableWidget::item:selected {
        background: rgba(59,130,246,0.3);
    }
    QHeaderView::section {
        background: #1e293b;
        color: #94a3b8;
        border: 1px solid rgba(148,163,184,0.2);
        padding: 4px;
        font-weight: 500;
    }
"""

_COMBO_STYLE = """
    QComboBox {
        background: #1e293b;
        color: #e5e7eb;
        border: 1px solid rgba(148,163,184,0.3);
        border-radius: 6px;
        padding: 4px 8px;
        min-width: 120px;
    }
    QComboBox:hover { border: 1px solid rgba(59,130,246,0.5); }
    QComboBox QAbstractItemView {
        background: #1e293b;
        color: #e5e7eb;
        selection-background-color: rgba(59,130,246,0.3);
    }
"""

_CHECKBOX_STYLE = """
    QCheckBox {
        color: #94a3b8;
        spacing: 4px;
    }
    QCheckBox::indicator {
        width: 14px;
        height: 14px;
        border: 1px solid rgba(148,163,184,0.4);
        border-radius: 3px;
        background: #1e293b;
    }
    QCheckBox::indicator:checked {
        background: rgba(59,130,246,0.6);
        border-color: rgba(59,130,246,0.8);
    }
"""

_ODS_NS = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}

_XLSX_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_XLSX_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_XLSX_OFFICEREL_NS = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)


def _read_csv_text(path: str) -> str:
    for enc in ("utf-8", "latin-1"):
        try:
            with open(path, encoding=enc, newline="") as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError(
        f"No se pudo decodificar el archivo «{os.path.basename(path)}».\n"
        "Se intentó con utf-8 y latin-1. Verifique la codificación del archivo."
    )


def _detect_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        counts = {d: sample.count(d) for d in (",", ";", "\t")}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else ","


def _parse_csv(path: str) -> list[list[str]]:
    text = _read_csv_text(path)
    delimiter = _detect_delimiter(text[:4096])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [row for row in reader]


def _parse_xlsx_openpyxl(path: str) -> dict[str, list[list[str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[list[str]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        sheets[name] = rows
    wb.close()
    return sheets


def _parse_xlsx_xml(path: str) -> dict[str, list[list[str]]]:
    sheets: dict[str, list[list[str]]] = {}

    with zipfile.ZipFile(path, "r") as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            ss_tree = ET.parse(zf.open("xl/sharedStrings.xml"))
            for si in ss_tree.iter(f"{{{_XLSX_NS}}}si"):
                parts: list[str] = []
                for t_elem in si.iter(f"{{{_XLSX_NS}}}t"):
                    parts.append(t_elem.text or "")
                shared.append("".join(parts))

        wb_tree = ET.parse(zf.open("xl/workbook.xml"))
        sheet_names: list[str] = []
        for s in wb_tree.iter(f"{{{_XLSX_NS}}}sheet"):
            sheet_names.append(s.get("name", "Hoja"))

        rid_map: dict[str, str] = {}
        rels_path = "xl/_rels/workbook.xml.rels"
        if rels_path in zf.namelist():
            rels_tree = ET.parse(zf.open(rels_path))
            for rel in rels_tree.iter(f"{{{_XLSX_REL_NS}}}Relationship"):
                rid_map[rel.get("Id", "")] = rel.get("Target", "")

        for idx, s_elem in enumerate(wb_tree.iter(f"{{{_XLSX_NS}}}sheet")):
            r_id = s_elem.get(
                f"{{{_XLSX_OFFICEREL_NS}}}id", ""
            )
            target = rid_map.get(r_id, f"worksheets/sheet{idx + 1}.xml")
            sheet_file = f"xl/{target}" if not target.startswith("/") else target.lstrip("/")
            name = sheet_names[idx] if idx < len(sheet_names) else f"Hoja {idx + 1}"

            if sheet_file not in zf.namelist():
                sheets[name] = []
                continue

            tree = ET.parse(zf.open(sheet_file))
            rows: list[list[str]] = []
            for row_elem in tree.iter(f"{{{_XLSX_NS}}}row"):
                cells: list[str] = []
                for c_elem in row_elem.iter(f"{{{_XLSX_NS}}}c"):
                    c_type = c_elem.get("t", "")
                    v_elem = c_elem.find(f"{{{_XLSX_NS}}}v")
                    val = v_elem.text if v_elem is not None else ""
                    if c_type == "s" and val:
                        idx_val = int(val)
                        val = shared[idx_val] if idx_val < len(shared) else val
                    cells.append(val or "")
                rows.append(cells)
            sheets[name] = rows

    return sheets


def _parse_xlsx(path: str) -> dict[str, list[list[str]]]:
    try:
        return _parse_xlsx_openpyxl(path)
    except ImportError:
        return _parse_xlsx_xml(path)


def _parse_ods(path: str) -> dict[str, list[list[str]]]:
    sheets: dict[str, list[list[str]]] = {}

    with zipfile.ZipFile(path, "r") as zf:
        tree = ET.parse(zf.open("content.xml"))

    root = tree.getroot()
    for tbl in root.iter(f"{{{_ODS_NS['table']}}}table"):
        name = tbl.get(f"{{{_ODS_NS['table']}}}name", "Hoja")
        rows: list[list[str]] = []
        for tr in tbl.iter(f"{{{_ODS_NS['table']}}}table-row"):
            cells: list[str] = []
            for tc in tr.iter(f"{{{_ODS_NS['table']}}}table-cell"):
                repeat = int(
                    tc.get(
                        f"{{{_ODS_NS['table']}}}number-columns-repeated", "1"
                    )
                )
                text_parts: list[str] = []
                for p in tc.iter(f"{{{_ODS_NS['text']}}}p"):
                    text_parts.append(p.text or "")
                cell_text = "\n".join(text_parts)
                cells.extend([cell_text] * repeat)
            rows.append(cells)

        while rows and all(c == "" for c in rows[-1]):
            rows.pop()
        if rows:
            max_col = max(
                (
                    max((i for i, c in enumerate(r) if c), default=-1) + 1
                    for r in rows
                ),
                default=0,
            )
            rows = [r[:max_col] for r in rows]

        sheets[name] = rows

    return sheets


class SpreadsheetViewer(QWidget):

    def __init__(self) -> None:
        super().__init__()
        self.current_path: str | None = None
        self._sheets: dict[str, list[list[str]]] = {}
        self._header_row = True
        self._build_ui()


    def _build_ui(self) -> None:
        toolbar = QHBoxLayout()
        toolbar.setContentsMargins(0, 0, 0, 0)
        toolbar.setSpacing(6)

        sheet_label = QLabel("📄 Hoja:")
        sheet_label.setStyleSheet("color: #94a3b8; font-size: 12px;")
        toolbar.addWidget(sheet_label)

        self.sheet_combo = QComboBox()
        self.sheet_combo.setStyleSheet(_COMBO_STYLE)
        self.sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)
        toolbar.addWidget(self.sheet_combo)

        self.header_check = QCheckBox("Primera fila como encabezado")
        self.header_check.setChecked(True)
        self.header_check.setStyleSheet(_CHECKBOX_STYLE)
        self.header_check.toggled.connect(self._on_header_toggled)
        toolbar.addWidget(self.header_check)

        toolbar.addStretch(1)

        self.info_label = QLabel()
        self.info_label.setStyleSheet("color: #94a3b8; font-size: 12px;")
        self.info_label.setAlignment(Qt.AlignVCenter | Qt.AlignRight)
        toolbar.addWidget(self.info_label)

        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectItems)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(_TABLE_STYLE)
        self.table.verticalHeader().setVisible(True)
        self.table.horizontalHeader().setStretchLastSection(True)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)
        layout.addLayout(toolbar)
        layout.addWidget(self.table, 1)


    def load_file(self, path: str) -> None:
        self.current_path = path
        self._sheets.clear()
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

        ext = Path(path).suffix.lower()

        try:
            if ext == ".csv":
                rows = _parse_csv(path)
                name = Path(path).stem
                self._sheets = {name: rows}
            elif ext in (".xlsx", ".xls"):
                self._sheets = self._load_xlsx(path)
            elif ext == ".ods":
                self._sheets = _parse_ods(path)
            else:
                self._show_error("Formato de archivo no reconocido.")
                return
        except zipfile.BadZipFile:
            self._show_error("El archivo está dañado o no es un formato válido.")
            return
        except Exception as exc:
            self._show_error(f"Error al leer el archivo:\n{exc}")
            return

        if not self._sheets:
            self._show_error("No se encontraron hojas en el archivo.")
            return

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        for name in self._sheets:
            self.sheet_combo.addItem(name)
        self.sheet_combo.blockSignals(False)

        self.sheet_combo.setVisible(len(self._sheets) > 1)

        self._display_sheet(list(self._sheets.keys())[0])

    def _load_xlsx(self, path: str) -> dict[str, list[list[str]]]:
        try:
            return _parse_xlsx(path)
        except Exception as exc:
            try:
                return _parse_xlsx_xml(path)
            except Exception:
                raise RuntimeError(
                    f"No se pudo leer el archivo.\n"
                    f"Instala openpyxl para mejor compatibilidad:\n"
                    f"  pip install openpyxl\n\n"
                    f"Error original: {exc}"
                ) from exc


    def _display_sheet(self, name: str) -> None:
        rows = self._sheets.get(name, [])
        self.table.clear()

        if not rows:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.info_label.setText("0 filas × 0 columnas")
            return

        num_cols = max(len(r) for r in rows)
        self._header_row = self.header_check.isChecked()

        if self._header_row and len(rows) > 0:
            headers = rows[0]
            data_rows = rows[1:]
            while len(headers) < num_cols:
                headers.append("")
            self.table.setColumnCount(num_cols)
            self.table.setHorizontalHeaderLabels(headers)
        else:
            data_rows = rows
            self.table.setColumnCount(num_cols)
            col_headers = [_col_letter(i) for i in range(num_cols)]
            self.table.setHorizontalHeaderLabels(col_headers)

        self.table.setRowCount(len(data_rows))

        for r_idx, row in enumerate(data_rows):
            for c_idx in range(num_cols):
                val = row[c_idx] if c_idx < len(row) else ""
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(r_idx, c_idx, item)

        header = self.table.horizontalHeader()
        if num_cols <= 20:
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
        else:
            header.setSectionResizeMode(QHeaderView.Interactive)
            self.table.resizeColumnsToContents()

        self.info_label.setText(
            f"{len(data_rows)} filas × {num_cols} columnas"
        )


    def _on_sheet_changed(self, index: int) -> None:
        if index < 0:
            return
        name = self.sheet_combo.itemText(index)
        if name in self._sheets:
            self._display_sheet(name)

    def _on_header_toggled(self, checked: bool) -> None:
        self._header_row = checked
        if self._sheets:
            name = self.sheet_combo.currentText()
            if name in self._sheets:
                self._display_sheet(name)


    def _show_error(self, msg: str) -> None:
        self.table.clear()
        self.table.setRowCount(1)
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels([""])
        item = QTableWidgetItem(f"⚠️ {msg}")
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(0, 0, item)
        self.info_label.setText("")
        QMessageBox.warning(self, "Error", msg)


def _col_letter(index: int) -> str:
    result = ""
    i = index
    while True:
        result = chr(ord("A") + i % 26) + result
        i = i // 26 - 1
        if i < 0:
            break
    return result
